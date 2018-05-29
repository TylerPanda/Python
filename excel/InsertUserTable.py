import os
from openpyxl import load_workbook

wb = load_workbook('excel_files/role.xlsx');
i = 0
for sheet in wb.sheetnames:
    print("\tCurrent Sheet is :" +  sheet);
    for row in wb[sheet].iter_rows(min_row = 2):
        single_row_values = []
        for cell in row:
            single_row_values.append(cell.value)
        # print(single_row_values)
        single_row_values.reverse()
        print("INSERT " + "INTO UserTable (UserID,UserName,Password,RealName,CreateUserID,UserStatus,UserAttrib,UserDepart,CreateTime,Comments,PID,UserCode,Sex,Email,Profession,TelOffice,TelHome,MobilePhone,Fax,Address,PageSize,BeginTime,LoginType) VALUES ({}, \"{}\", \"{}\", \"{}\", {}, {}, {}, {}, {}, \"{}\", \"{}\", {}, {}, \"{}\", {}, {}, {}, {}, {}, {}, {}, {}, \"{}\" )".format(single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop(), single_row_values.pop()))
