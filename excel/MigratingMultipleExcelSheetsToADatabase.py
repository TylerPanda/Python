import os
from openpyxl import load_workbook


big_list_of_all_rows = []
for file in os.listdir('excel_files'):
    print('Current file is ' + file)
    wb = load_workbook(os.path.join('excel_files', file))

    for sheetname in wb.sheetnames:
        print('\tCurrent sheet is ' + sheetname)
        ws = wb[sheetname]

        for row in ws.iter_rows(min_row = 2):
            single_row_values = []

            for cell in row:
                single_row_values.append(cell.value)
            print(single_row_values)
            big_list_of_all_rows.append(single_row_values)
print(big_list_of_all_rows)
