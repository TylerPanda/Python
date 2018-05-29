from openpyxl import load_workbook

wb = load_workbook('excel_files/revenue.xlsx')

# Show the sheet names
sheets = wb.sheetnames
print(sheets)

# Print the cell value
for sheet in sheets:
    for row in wb[sheet]:
        for cell in row:
            print(cell.value)

# Show the data by rows
for sheet in sheets:
    for row in wb[sheet].rows:
        for cell in row:
            print(cell.value)

# Show the data by column
for sheet in sheets:
    for column in wb[sheet].columns:
        for cell in column:
            print(cell.value)
